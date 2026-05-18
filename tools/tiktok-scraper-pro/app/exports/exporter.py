"""
Export Engine — CSV / JSON / Excel output
"""
import csv
import json
import os
from datetime import datetime
from typing import Any


def _timestamp_filename(prefix: str, ext: str) -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{ts}.{ext}"


def export_csv(rows: list[dict], output_dir: str, prefix: str = "export") -> str:
    """Write rows to a CSV file. Returns the absolute file path."""
    os.makedirs(output_dir, exist_ok=True)
    path = os.path.join(output_dir, _timestamp_filename(prefix, "csv"))
    if not rows:
        return path
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)
    return path


def export_json(rows: list[dict], output_dir: str, prefix: str = "export") -> str:
    """Write rows to a JSON file. Returns the absolute file path."""
    os.makedirs(output_dir, exist_ok=True)
    path = os.path.join(output_dir, _timestamp_filename(prefix, "json"))
    with open(path, "w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)
    return path


def export_excel(rows: list[dict], output_dir: str, prefix: str = "export") -> str:
    """Write rows to an Excel .xlsx file using openpyxl. Returns the absolute file path."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    os.makedirs(output_dir, exist_ok=True)
    path = os.path.join(output_dir, _timestamp_filename(prefix, "xlsx"))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    if not rows:
        wb.save(path)
        return path

    # Header row
    headers = list(rows[0].keys())
    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(bold=True, color="38BDF8")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h.replace("_", " ").title())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        for col, key in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col, value=row.get(key, ""))
        if row_idx % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = PatternFill("solid", fgColor="0F172A")

    # Auto column widths
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    wb.save(path)
    return path


def export(
    rows: list[dict],
    output_dir: str,
    fmt: str,
    prefix: str = "tiktok_results",
) -> str:
    """Dispatch to the correct exporter. fmt: 'csv' | 'json' | 'excel'"""
    if fmt == "csv":
        return export_csv(rows, output_dir, prefix)
    elif fmt == "json":
        return export_json(rows, output_dir, prefix)
    elif fmt == "excel":
        return export_excel(rows, output_dir, prefix)
    raise ValueError(f"Unknown export format: {fmt}")
